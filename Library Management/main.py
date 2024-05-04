# u r in showimage command part dont forget dumbie
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"


root=Tk()
root.title("Library Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)




file=pathlib.Path('Student_Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No" 
    sheet['B1']="Name" 
    sheet['C1']="Class" 
    sheet['D1']="Gender" 
    sheet['E1']="DOB" 
    sheet['F1']="Date Of Registration" 
    sheet['G1']="Roll No" 
    sheet['H1']="Address" 
    sheet['I1']="Stream" 
    sheet['J1']="Phone No"
    sheet['K1']="Book" 
    sheet['L1']="Section" 

    file.save('Student_Data.xlsx')

#exit
def Exit():
    root.destroy()
########################3how image

def showimage():
    global filename
    global img
    # Determine the appropriate file dialog based on the operating system
    if os.name == "posix":  # Linux
        filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                               title="Select Image File",
                                               filetypes=(("Image Files", "*.jpg;*.png"), ("All Files", "*.*")))
    elif os.name == "nt":  # Windows
        filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                               title="Select Image File",
                                               filetypes=(("JPG File", "*.jpg"),
                                                         ("PNG File", "*.png"),
                                                         ("All Files", "*.*")))
    else:
        messagebox.showerror("Error", "Unsupported operating system")
        return
    
    if filename:  # Check if a file is selected
        img = Image.open(filename)
        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        
        # Configure the label to display the selected image
        lbl.config(image=photo2)
        lbl.image = photo2

####################registration no
def registration_no():
    file=openpyxl.load_workbook('Student_Data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
   

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")

###########clear###################3
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Class.set("Select Class")
    Section.set('')
    Address.set('')
    Roll.set('')
    Book.set('')
    Phone.set('')
    Stream.set('')


    registration_no()

    saveButton.config(state= 'normal')

    img1=PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

############SAVE##############
def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Seems Like you forget to Select Gender!")

    D2=DOB.get()
    D1=Date.get()
    Rol=Roll.get()
    A1=Address.get()
    S1=Stream.get()
    P1=Phone.get()
    B1=Book.get()
    Sec=Section.get()

    if N1=="" or C1=="Select Class" or D2=="" or Rol=="" or S1=="" or P1=="" or Sec=="" or A1=="" or B1=="":
     messagebox.showerror("error","Something is missing!")

    else:
        file=openpyxl.load_workbook('Student_Data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Rol)
        sheet.cell(column=8,row=sheet.max_row,value=A1)
        sheet.cell(column=9,row=sheet.max_row,value=S1)
        sheet.cell(column=10,row=sheet.max_row,value=P1)
        sheet.cell(column=11,row=sheet.max_row,value=B1)
        sheet.cell(column=12,row=sheet.max_row,value=Sec)
        file.save(r'Student_Data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!")
        messagebox.showinfo("info","Data Successfully Stored!!!")

        Clear() #clear entry box

        registration_no() #recheck reg no and reissue new no

############search##############

def search():
    text = Search.get()  # Taking input

    # Validate input to ensure only numeric characters are entered
    if not text.isdigit():
        messagebox.showerror("Invalid", "Please enter a valid registration number.")
        return

    Clear()  # Clear all data already available in entry box and other
    saveButton.config(state='disable')  # After clicking on search, save button will be disabled

    file = openpyxl.load_workbook("Student_Data.xlsx")
    sheet = file.active

    registration_found = False  # Flag to track if the registration number is found

    for row in sheet.rows:
        if row[0].value == int(text):
            reg_number = row[0].row  # Extract the row number directly
            registration_found = True
            break  # Stop searching once the registration number is found

    if not registration_found:
        messagebox.showerror("Invalid", "Registration number not found.")
        return

    # Use the extracted row number to fetch data from the corresponding row
    x1 = sheet.cell(row=reg_number, column=1).value
    x2 = sheet.cell(row=reg_number, column=2).value
    x3 = sheet.cell(row=reg_number, column=3).value
    x4 = sheet.cell(row=reg_number, column=4).value
    x5 = sheet.cell(row=reg_number, column=5).value
    x6 = sheet.cell(row=reg_number, column=6).value
    x7 = sheet.cell(row=reg_number, column=7).value
    x8 = sheet.cell(row=reg_number, column=8).value
    x9 = sheet.cell(row=reg_number, column=9).value
    x10 = sheet.cell(row=reg_number, column=10).value
    x11 = sheet.cell(row=reg_number, column=11).value
    x12 = sheet.cell(row=reg_number, column=12).value

    # Update entry fields with fetched data
    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4 == 'Female':
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Roll.set(x7)
    Address.set(x8)
    Stream.set(x9)
    Phone.set(x10)
    Book.set(x11)
    Section.set(x12)

    # Load and display the corresponding image
    img = Image.open("Student Images/" + str(x1) + ".jpg")
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

##########update
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Rol = Roll.get()
    A1 = Address.get()
    S1 = Stream.get()
    P1 = Phone.get()
    B1 = Book.get()
    Sec = Section.get()

    file = openpyxl.load_workbook("Student_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            reg_number = row[0].row  # Extract the row number directly

            sheet.cell(column=2, row=reg_number, value=N1)
            sheet.cell(column=3, row=reg_number, value=C1)
            sheet.cell(column=4, row=reg_number, value=G1)
            sheet.cell(column=5, row=reg_number, value=D2)
            sheet.cell(column=6, row=reg_number, value=D1)
            sheet.cell(column=7, row=reg_number, value=Rol)
            sheet.cell(column=8, row=reg_number, value=A1)
            sheet.cell(column=9, row=reg_number, value=S1)
            sheet.cell(column=10, row=reg_number, value=P1)
            sheet.cell(column=11, row=reg_number, value=B1)
            sheet.cell(column=12, row=reg_number, value=Sec)

            break  # Stop the loop once the registration number is found

    file.save(r'Student_Data.xlsx')

    try:
        img.save("Student Images/" + str(R1) + ".jpg")
    except:
        pass
    messagebox.showinfo("Update", "Updated Successfully!!!")

    Clear()







#gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
       
    else:
        gender="Female"
       




#top frames
Label(root,text="Deepshikha College",width=10,height=3,bg="#f0687c",font='arial 15 bold').pack(side=TOP,fill=X)
Label(root,text="LIBRARY",width=10,height=2,bg="#c36464",fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)

#search box
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
imageicon3=PhotoImage(file="Images/search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg='#68ddfa',font='arial 13 bold',command=search)
Srch.place(x=1060,y=66)

imageicon4=PhotoImage(file="Images/Layer 4.png")
Update_button=Button(root,image=imageicon4,bg="#c36464",command=Update)
Update_button.place(x=110,y=64)

#Registration and Date
Label(root,text="Registration No:", font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()


today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)


#student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date Of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)


Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Section:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Roll No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name = StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB = StringVar()
dob_entry = Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)


radio= IntVar()
R1 = Radiobutton(obj,text="M", variable=radio, value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj,text="F", variable=radio, value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)


Section = StringVar()
section_entry = Entry(obj,textvariable=Section,width=20,font="arial 10")
section_entry.place(x=630,y=100)

Roll = StringVar()
roll_entry = Entry(obj,textvariable=Roll,width=20,font="arial 10")
roll_entry.place(x=630,y=150)

Class= Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")


#other
obj2=LabelFrame(root,text="Other Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Book:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

Book= StringVar()
book_entry = Entry(obj2,textvariable=Book,width=20,font="arial 10")
book_entry.place(x=160,y=50)

Phone= StringVar()
phone_entry = Entry(obj2,textvariable=Phone,width=20,font="arial 10")
phone_entry.place(x=160,y=100)



Label(obj2,text="Stream:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Address:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

Stream= StringVar()
stream_entry = Entry(obj2,textvariable=Stream,width=20,font="arial 10")
stream_entry.place(x=630,y=50)

Address= StringVar()
address_entry = Entry(obj2,textvariable=Address,width=20,font="arial 10")
address_entry.place(x=630,y=100)


#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button

Button(root,text="Upload",width=19,height=2,font="airal 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="airal 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="airal 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="airal 12 bold",bg="grey",command=Exit).place(x=1000,y=610)






root.mainloop()